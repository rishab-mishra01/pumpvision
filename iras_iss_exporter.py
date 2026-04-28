"""
IRAS ISS Automated Exporter
============================
Two modes:

  "boundary" — for each date in the list, search backwards from 06:00
               collecting the last Totalizer End per nozzle independently.
               Continues until all 6 nozzles (7, 11, 15, 16, 17, 18) are
               found or 48 windows have been checked.
               Returns {nozzle_no: totalizer_end_value} for the operational
               day boundary (i.e. the 6AM opening totalizer per nozzle).

  "full"     — export every 30-min window of the full 24-hour shift
               (06:00 on START_DATE  →  06:00 the next calendar day).

USAGE:
    python -X utf8 iras_iss_exporter.py
"""

import asyncio
import io
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────

IRAS_URL   = "https://iras.iocliras.in/login"
USERNAME   = "206858"
PASSWORD   = "Shree@26"

OUTPUT_FOLDER = r"C:\IRAS_Data\ISS"

# Date(s) to process — one entry per shift day.
# A "shift day" runs from 06:00 on this date to 06:00 the next calendar day.
SHIFT_DATES = ["2026-02-26", "2026-02-27"]

SHIFT_START_HR = 6   # shifts start at 6 AM

# "boundary" → find last transaction before 6am (backwards search, one file per date)
# "full"     → export all 48 half-hour windows of the shift
RUN_MODE = "boundary"

TABLE_LOAD_TIMEOUT = 30   # seconds to wait for table after clicking Show
DOWNLOAD_TIMEOUT   = 30   # seconds to wait for Excel download
DELAY_BETWEEN      = 2    # seconds between exports

# All active nozzles at the outlet
NOZZLES = {7, 11, 15, 16, 17, 18}


# ─────────────────────────────────────────────
# EXCEL PARSER
# ─────────────────────────────────────────────

def parse_totalizer_ends(filepath: Path) -> dict:
    """
    Parse an ISS Excel file and return {nozzle_no: totalizer_end} reflecting
    the LAST transaction row per nozzle found in the file.

    Locates the header row by searching for a row that contains both 'nozzle'
    and 'totalizer' (case-insensitive), then reads all data rows below it.
    The last row for each nozzle wins (rows are chronological in IRAS exports).
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"    [parse error] Could not open {filepath.name}: {e}")
        return {}

    # Prefer a sheet whose name contains 'iss' or 'issue'
    sheet_name = next(
        (n for n in wb.sheetnames if "iss" in n.lower() or "issue" in n.lower()),
        wb.sheetnames[0],
    )
    ws = wb[sheet_name]

    # Find header row
    headers = None
    header_row_idx = 0
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        joined = " ".join(cells)
        if "nozzle" in joined and "totalizer" in joined:
            headers = cells
            header_row_idx = row_idx
            break

    if headers is None:
        wb.close()
        print(f"    [parse warn] Header row not found in {filepath.name}")
        return {}

    # Locate required columns
    nozzle_col = next(
        (i for i, h in enumerate(headers) if "nozzle" in h and "no" in h),
        next((i for i, h in enumerate(headers) if "nozzle" in h), None),
    )
    tot_end_col = next(
        (i for i, h in enumerate(headers) if "totalizer" in h and "end" in h),
        None,
    )

    if nozzle_col is None or tot_end_col is None:
        wb.close()
        print(f"    [parse warn] Could not locate Nozzle/Totalizer End columns in {filepath.name}")
        return {}

    # Collect last Totalizer End per nozzle (last row wins = latest transaction)
    result = {}
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        raw_nozzle  = row[nozzle_col]
        raw_tot_end = row[tot_end_col]
        if raw_nozzle is None or raw_tot_end is None:
            continue
        try:
            nozzle_int = int(raw_nozzle)
        except (ValueError, TypeError):
            continue
        if nozzle_int in NOZZLES:
            result[nozzle_int] = raw_tot_end   # overwrite → last row wins

    wb.close()
    return result


# ─────────────────────────────────────────────
# WINDOW GENERATORS
# ─────────────────────────────────────────────

def shift_windows(shift_date: str, start_hr: int = 6):
    """
    Return 48 half-hour windows covering one full 24-hour shift.
    Each entry: (from_date, from_h, from_m, to_date, to_h, to_m)
    """
    base = datetime.strptime(shift_date, "%Y-%m-%d").replace(hour=start_hr, minute=0)
    windows = []
    t = base
    for _ in range(48):
        t_end = t + timedelta(minutes=30)
        windows.append((
            t.strftime("%Y-%m-%d"),   t.hour,   t.minute,
            t_end.strftime("%Y-%m-%d"), t_end.hour, t_end.minute,
        ))
        t = t_end
    return windows


def boundary_windows(boundary_date: str, boundary_hr: int = 6, max_steps: int = 48):
    """
    Return windows starting at boundary_hr:00, stepping backwards 30 min each time.
      step 0 → boundary_hr:00 – boundary_hr:30   (first window OF the shift)
      step 1 → (boundary_hr - 0:30) – boundary_hr:00
      step 2 → (boundary_hr - 1:00) – (boundary_hr - 0:30)
      ...
    Stops after max_steps (default = 48 = 24 hours back).
    Each entry: (from_date, from_h, from_m, to_date, to_h, to_m)
    """
    base = datetime.strptime(boundary_date, "%Y-%m-%d").replace(hour=boundary_hr, minute=0)
    windows = []
    for step in range(max_steps):
        t_start = base - timedelta(minutes=30 * step)
        t_end   = t_start + timedelta(minutes=30)
        windows.append((
            t_start.strftime("%Y-%m-%d"), t_start.hour, t_start.minute,
            t_end.strftime("%Y-%m-%d"),   t_end.hour,   t_end.minute,
        ))
    return windows


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def filename_safe(fd, fh, fm, td, th, tm):
    return f"ISS_{fd}_{fh:02d}{fm:02d}_{td}_{th:02d}{tm:02d}.xlsx"


def build_date_str(date_str: str, hour: int, minute: int) -> str:
    """Format a date+time as IRAS expects: DD-MM-YYYY hh:mm:ss am/pm (12-hour)."""
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    am_pm = "am" if hour < 12 else "pm"
    h12   = hour % 12 or 12
    return f"{dt.day:02d}-{dt.month:02d}-{dt.year} {h12:02d}:{minute:02d}:00 {am_pm}"


# ─────────────────────────────────────────────
# DATE / TIME FIELD
# ─────────────────────────────────────────────

async def set_datetime_field(page, field_label: str, date_str: str, hour: int, minute: int):
    """
    Set a From Date or To Date field.
    Uses fill() for React state updates; falls back to the React native setter.
    """
    value = build_date_str(date_str, hour, minute)

    # Two date inputs share the same placeholder — From is .first, To is .last
    selector = "input[placeholder*='DD-MM-YYYY']"
    field = page.locator(selector).first if "From" in field_label else page.locator(selector).last

    await field.wait_for(state="visible", timeout=5_000)
    await field.fill(value)
    await page.wait_for_timeout(200)

    # Verify React accepted it; fall back to native setter if not
    actual = await field.input_value()
    if actual != value:
        el = await field.element_handle()
        await page.evaluate(
            "([el, val]) => {"
            "  const setter = Object.getOwnPropertyDescriptor("
            "    window.HTMLInputElement.prototype, 'value').set;"
            "  setter.call(el, val);"
            "  el.dispatchEvent(new Event('input',  {bubbles:true}));"
            "  el.dispatchEvent(new Event('change', {bubbles:true}));"
            "}",
            [el, value]
        )
        await page.wait_for_timeout(200)
        actual = await field.input_value()

    await page.keyboard.press("Tab")
    await page.wait_for_timeout(200)
    print(f"    {field_label}: {actual!r}")


# ─────────────────────────────────────────────
# NAVIGATION
# ─────────────────────────────────────────────

async def set_as_per_actual(page):
    """Select 'Actual Transaction Date and Time' in the As Per MUI Select."""
    as_per = page.locator("div[role='combobox'][aria-labelledby*='As Per']")
    await as_per.wait_for(state="visible", timeout=5_000)
    await as_per.click()
    await page.wait_for_timeout(600)
    option = page.locator("li[role='option']").filter(has_text="Actual")
    await option.first.wait_for(state="visible", timeout=5_000)
    await option.first.click()
    await page.wait_for_timeout(400)
    print("  [OK] As Per = Actual Transaction Date and Time")


async def navigate_to_iss(page, shift_date: str = SHIFT_DATES[0]):
    """
    1. Click FCC Data in the left nav
    2. Click the ... overflow tab
    3. Click Issue(ISS) in the popup
    4. If shift_date is older than 7 days, click the archive toggle
    5. Set As Per = Actual Transaction Date and Time
    """
    await page.locator("text=FCC Data").click()
    await page.wait_for_timeout(2000)

    overflow = page.locator("button[role='tab']:has-text('...')")
    await overflow.wait_for(state="visible", timeout=10_000)
    await overflow.click()
    await page.wait_for_timeout(1000)

    iss = page.locator("li.app-tab-list:has-text('Issue(ISS)')")
    await iss.wait_for(state="visible", timeout=5_000)
    await iss.click()
    await page.wait_for_timeout(1500)

    # Click archive toggle only when date is more than 7 days in the past
    days_old = (datetime.now() - datetime.strptime(shift_date, "%Y-%m-%d")).days
    if days_old > 7:
        toggle = page.locator(".MuiFormControlLabel-root:has-text('Last 7 days Report')")
        await toggle.wait_for(state="visible", timeout=10_000)
        await toggle.click()
        await page.wait_for_timeout(600)
        print("  [OK] Archive toggle enabled (date is >7 days old)")
    else:
        print("  [--] Date within last 7 days — archive toggle not needed")

    await set_as_per_actual(page)
    print("[OK] Navigated to FCC Data > Issue(ISS)")


# ─────────────────────────────────────────────
# EXPORT ONE WINDOW
# ─────────────────────────────────────────────

async def export_window(page, output_dir: Path,
                        fd: str, fh: int, fm: int,
                        td: str, th: int, tm: int) -> bool | None:
    """
    Set the date range, click Show, and download the Excel file.

    Returns:
        True  — file downloaded (table had data)
        None  — table was empty (no records for this window, not an error)
        False — an error occurred
    """
    fname = filename_safe(fd, fh, fm, td, th, tm)
    fpath = output_dir / fname

    if fpath.exists():
        print(f"  [skip] Already exists: {fname}")
        return True

    try:
        await set_datetime_field(page, "From Date", fd, fh, fm)
        await set_datetime_field(page, "To Date",   td, th, tm)

        await page.locator("button:has-text('Show')").first.click()
        await page.wait_for_timeout(1500)

        # Wait for table to finish loading (rows present OR no-data overlay)
        try:
            await page.wait_for_selector(
                ".ag-row, .ag-overlay-no-rows-wrapper",
                timeout=TABLE_LOAD_TIMEOUT * 1000
            )
        except PlaywrightTimeout:
            pass

        await page.wait_for_timeout(500)

        row_count = await page.locator(".ag-row").count()
        if row_count == 0:
            print(f"  [empty] No data")
            return None   # not a failure

        # Download Excel
        async with page.expect_download(timeout=DOWNLOAD_TIMEOUT * 1000) as dl_info:
            await page.locator("button.export-excel-button").click()

        download = await dl_info.value
        await download.save_as(str(fpath))
        print(f"  [saved] {fname}  ({row_count} rows)")
        return True

    except PlaywrightTimeout as e:
        print(f"  [timeout] {fname}: {e}")
        return False
    except Exception as e:
        print(f"  [error] {fname}: {e}")
        return False


# ─────────────────────────────────────────────
# SESSION KEEPALIVE
# ─────────────────────────────────────────────

async def handle_session_expiry(page):
    try:
        if await page.locator("text=session").first.is_visible(timeout=1000):
            for btn in ["OK", "Continue", "Stay", "Yes"]:
                b = page.locator(f"button:has-text('{btn}')").first
                if await b.is_visible(timeout=500):
                    await b.click()
                    await page.wait_for_timeout(1000)
                    return True
    except Exception:
        pass
    return False


# ─────────────────────────────────────────────
# BOUNDARY MODE
# ─────────────────────────────────────────────

async def run_boundary(page, output_dir: Path, shift_date: str) -> dict:
    """
    Search backwards from 06:00 on shift_date collecting the last Totalizer End
    per nozzle independently across multiple 30-min windows.

    Each downloaded window is parsed immediately; nozzles are removed from the
    'remaining' set as they are found.  The search stops when all 6 nozzles are
    resolved or 48 windows have been checked (24 hours back).

    Windows start at step 1 (05:30-06:00) so we never include transactions from
    the new shift (step 0 = 06:00-06:30 belongs to the day being opened, not the
    day being closed).

    Returns {nozzle_no: totalizer_end_value} for every nozzle found.
    """
    print(f"\n[boundary] {shift_date}: collecting last Totalizer End per nozzle before 06:00")
    wins = boundary_windows(shift_date, SHIFT_START_HR)

    found: dict[int, float] = {}
    remaining = set(NOZZLES)

    for (fd, fh, fm, td, th, tm) in wins[1:]:   # skip step-0 (06:00-06:30 = new shift)
        if not remaining:
            break

        print(f"  [{shift_date}] checking {fd} {fh:02d}:{fm:02d}-{td} {th:02d}:{tm:02d}  "
              f"(still need nozzles: {sorted(remaining)})")

        await handle_session_expiry(page)
        result = await export_window(page, output_dir, fd, fh, fm, td, th, tm)

        if result is True:
            fpath = output_dir / filename_safe(fd, fh, fm, td, th, tm)
            window_data = parse_totalizer_ends(fpath)
            for nozzle, tot_end in window_data.items():
                if nozzle in remaining:
                    found[nozzle] = tot_end
                    remaining.discard(nozzle)
                    print(f"    [+] Nozzle {nozzle}: Totalizer End = {tot_end}")

        await asyncio.sleep(DELAY_BETWEEN)

    if remaining:
        print(f"  [WARN] No data found for nozzles: {sorted(remaining)}")
    print(f"\n  [boundary result] {shift_date}: {found}")
    return found


# ─────────────────────────────────────────────
# FULL SHIFT MODE
# ─────────────────────────────────────────────

async def run_full_shift(page, output_dir: Path, shift_date: str):
    """Export all 48 half-hour windows of the 06:00-to-06:00 shift."""
    windows = shift_windows(shift_date, SHIFT_START_HR)
    total = len(windows)
    success = skipped = failed = 0

    print(f"\n[full shift] {shift_date}  06:00 -> next day 06:00  ({total} windows)")

    for (fd, fh, fm, td, th, tm) in windows:
        print(f"  {fd} {fh:02d}:{fm:02d} - {td} {th:02d}:{tm:02d}", end="  ")
        await handle_session_expiry(page)
        result = await export_window(page, output_dir, fd, fh, fm, td, th, tm)
        if result is True:
            success += 1
        elif result is None:
            skipped += 1
        else:
            failed += 1
        await asyncio.sleep(DELAY_BETWEEN)

    print(f"\n  Shift {shift_date}: {success} downloaded, {skipped} empty, {failed} errors")
    return success, skipped, failed


# ─────────────────────────────────────────────
# LOGIN WAIT
# ─────────────────────────────────────────────

async def auto_fill_login(page):
    """
    1. Select 'Dealer' from the role dropdown.
    2. Fill Username and Password.
    Leaves CAPTCHA for manual entry.
    """
    try:
        # Wait for role dropdown — could be a <select> or a custom MUI select
        await page.wait_for_timeout(1000)

        # Try native <select> first
        sel = page.locator("select").first
        if await sel.count() > 0 and await sel.is_visible(timeout=2000):
            await sel.select_option(label="Dealer")
            print("[OK] Selected Dealer via <select>")
        else:
            # MUI-style dropdown — click to open, then pick option
            dropdown = page.locator("div[role='combobox'], .MuiSelect-select").first
            await dropdown.wait_for(state="visible", timeout=5000)
            await dropdown.click()
            await page.wait_for_timeout(600)
            dealer_opt = page.locator("li[role='option']:has-text('Dealer'), option:has-text('Dealer')").first
            await dealer_opt.wait_for(state="visible", timeout=5000)
            await dealer_opt.click()
            print("[OK] Selected Dealer via MUI dropdown")

        await page.wait_for_timeout(800)

        # Fill Username (206858) — appears after Dealer is selected
        for sel in ["input[name='username']", "input[name='userId']",
                    "input[placeholder*='Username']", "input[placeholder*='username']",
                    "input[placeholder*='User']"]:
            un = page.locator(sel).first
            if await un.count() > 0 and await un.is_visible(timeout=1000):
                await un.fill(USERNAME)
                print(f"[OK] Username filled ({USERNAME})")
                break

        # Fill Password
        pw = page.locator("input[type='password']").first
        if await pw.count() > 0 and await pw.is_visible(timeout=2000):
            await pw.fill(PASSWORD)
            print("[OK] Password filled")

    except Exception as e:
        print(f"[--] Auto-fill failed: {e}")
        print("     Please fill in Dealer / Username / Password manually.")


async def wait_for_login(page):
    await auto_fill_login(page)

    print()
    print("=" * 55)
    print("  ACTION REQUIRED")
    print("=" * 55)
    print("  Fill in the CAPTCHA and click Login")
    print("  (script waits up to 5 minutes)")
    print("=" * 55)
    print()
    await page.wait_for_function(
        "() => !window.location.href.includes('/login')",
        timeout=300_000
    )
    print("[OK] Login detected")
    await page.wait_for_timeout(2000)


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────

async def main():
    output_dir = Path(OUTPUT_FOLDER)
    output_dir.mkdir(parents=True, exist_ok=True)

    print()
    print("IRAS ISS Exporter")
    print(f"Mode        : {RUN_MODE}")
    print(f"Shift dates : {SHIFT_DATES}")
    print(f"Output      : {OUTPUT_FOLDER}")
    print()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, args=["--start-maximized"])
        context = await browser.new_context(
            accept_downloads=True,
            viewport={"width": 1400, "height": 900},
        )
        page = await context.new_page()

        # Login
        await page.goto(IRAS_URL, wait_until="networkidle")
        await page.wait_for_timeout(1500)

        await wait_for_login(page)

        # Navigate once to ISS (use first shift date for the archive toggle decision)
        await navigate_to_iss(page, shift_date=SHIFT_DATES[0])

        # Run selected mode for each shift date
        for shift_date in SHIFT_DATES:
            if RUN_MODE == "boundary":
                totalizers = await run_boundary(page, output_dir, shift_date)
                if totalizers:
                    print(f"\n  --> 6AM opening totalizers for {shift_date}:")
                    for nozzle in sorted(totalizers):
                        print(f"       Nozzle {nozzle}: {totalizers[nozzle]}")
                else:
                    print(f"\n  --> [WARN] No totalizer data found for {shift_date}")
            elif RUN_MODE == "full":
                await run_full_shift(page, output_dir, shift_date)

        print("\n[DONE]")
        await browser.close()


if __name__ == "__main__":
    asyncio.run(main())
