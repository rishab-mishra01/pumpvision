"""
IRAS ISS Automated Exporter
============================
Two modes:

  "boundary" — XG (nozzle 11) pre-check via Shift Totalizer first, then
               ISS backward search for the 5 active nozzles (7, 15, 16, 17, 18).
               If XG movement on the Shift Totalizer exceeds 7L a real sale
               occurred and nozzle 11 is added back into the ISS search.
               Returns {nozzle_no: totalizer_end_value} for all 6 nozzles.

  "full"     — export every 30-min window of the full 24-hour shift
               (06:00 on START_DATE  →  06:00 the next calendar day).

USAGE:
    python -X utf8 iras_iss_exporter.py
"""

import asyncio
import io
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout
from iras_proxy import iras_proxy_cfg, IRAS_PROXY_ENABLED, safe_exc_name

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# Add project root to path so we can import the Flask app and models
sys.path.insert(0, str(Path(__file__).parent.parent))

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────

IRAS_URL   = "https://iras.iocliras.in/login"
USERNAME   = "206858"
PASSWORD   = "Shree@26"

OUTPUT_FOLDER = r"C:\IRAS_Data\ISS"
SHIFT_TOTALIZER_FOLDER = r"C:\IRAS_Data\ShiftTotalizer"

# Date(s) to process — one entry per shift day.
# A "shift day" runs from 06:00 on this date to 06:00 the next calendar day.
SHIFT_DATES = [
    "2026-04-12", "2026-04-13", "2026-04-14", "2026-04-15",
    "2026-04-16", "2026-04-17", "2026-04-18",
]

SHIFT_START_HR = 6   # shifts start at 6 AM

# "boundary" → find last transaction before 6am (backwards search, one file per date)
# "full"     → export all 48 half-hour windows of the shift
RUN_MODE = "pump_test_scan"

TABLE_LOAD_TIMEOUT = 30   # seconds to wait for table after clicking Show
DOWNLOAD_TIMEOUT   = 30   # seconds to wait for Excel download
DELAY_BETWEEN      = 2    # seconds between exports

# All nozzles at the outlet
NOZZLES = {7, 11, 15, 16, 17, 18}

# The 5 nozzles included in the ISS backward search.
# Nozzle 11 (XG) is resolved via the Shift Totalizer pre-check before the ISS search runs,
# so it is excluded here unless the pre-check falls back to ISS.
ACTIVE_NOZZLES = {7, 15, 16, 17, 18}

XG_NOZZLE = 11
XG_MOVEMENT_THRESHOLD = 7.0   # 5L pump test + 2L buffer; movement above this = genuine sale

# ─── CLI argument override (used when launched from the web app) ──────────
import argparse as _argparse
_parser = _argparse.ArgumentParser(add_help=False)
_parser.add_argument("--dates", nargs="+")
_parser.add_argument("--mode", default=None)
_cli, _ = _parser.parse_known_args()
if _cli.dates:
    SHIFT_DATES = _cli.dates
if _cli.mode:
    RUN_MODE = _cli.mode


# ─────────────────────────────────────────────
# EXCEL PARSER
# ─────────────────────────────────────────────

def parse_totalizer_ends(filepath: Path) -> dict:
    """
    Parse an ISS Excel file and return {nozzle_no: totalizer_end} reflecting
    the LAST transaction row per nozzle found in the file.

    Locates the header row by searching for a row that contains both 'nozzle'
    and 'totalizer' (case-insensitive), then reads all data rows below it.
    The last row for each nozzle wins (rows are chronological in IRAS exports).
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"    [parse error] Could not open {filepath.name}: {e}")
        return {}

    # Prefer a sheet whose name contains 'iss' or 'issue'
    sheet_name = next(
        (n for n in wb.sheetnames if "iss" in n.lower() or "issue" in n.lower()),
        wb.sheetnames[0],
    )
    ws = wb[sheet_name]

    # Find header row
    headers = None
    header_row_idx = 0
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        joined = " ".join(cells)
        if "nozzle" in joined and "totalizer" in joined:
            headers = cells
            header_row_idx = row_idx
            break

    if headers is None:
        wb.close()
        print(f"    [parse warn] Header row not found in {filepath.name}")
        return {}

    # Locate required columns
    nozzle_col = next(
        (i for i, h in enumerate(headers) if "nozzle" in h and "no" in h),
        next((i for i, h in enumerate(headers) if "nozzle" in h), None),
    )
    tot_end_col = next(
        (i for i, h in enumerate(headers) if "totalizer" in h and "end" in h),
        None,
    )

    if nozzle_col is None or tot_end_col is None:
        wb.close()
        print(f"    [parse warn] Could not locate Nozzle/Totalizer End columns in {filepath.name}")
        return {}

    # Collect last Totalizer End per nozzle (last row wins = latest transaction)
    result = {}
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        raw_nozzle  = row[nozzle_col]
        raw_tot_end = row[tot_end_col]
        if raw_nozzle is None or raw_tot_end is None:
            continue
        try:
            nozzle_int = int(raw_nozzle)
        except (ValueError, TypeError):
            continue
        if nozzle_int in NOZZLES:
            result[nozzle_int] = raw_tot_end   # overwrite → last row wins

    wb.close()
    return result


# ─────────────────────────────────────────────
# WINDOW GENERATORS
# ─────────────────────────────────────────────

def shift_windows(shift_date: str, start_hr: int = 6):
    """
    Return 48 half-hour windows covering one full 24-hour shift.
    Each entry: (from_date, from_h, from_m, to_date, to_h, to_m)
    """
    base = datetime.strptime(shift_date, "%Y-%m-%d").replace(hour=start_hr, minute=0)
    windows = []
    t = base
    for _ in range(48):
        t_end = t + timedelta(minutes=30)
        windows.append((
            t.strftime("%Y-%m-%d"),   t.hour,   t.minute,
            t_end.strftime("%Y-%m-%d"), t_end.hour, t_end.minute,
        ))
        t = t_end
    return windows


def boundary_windows(boundary_date: str, boundary_hr: int = 6, max_steps: int = 48):
    """
    Return windows starting at boundary_hr:00, stepping backwards 30 min each time.
      step 0 → boundary_hr:00 – boundary_hr:30   (first window OF the shift)
      step 1 → (boundary_hr - 0:30) – boundary_hr:00
      step 2 → (boundary_hr - 1:00) – (boundary_hr - 0:30)
      ...
    Stops after max_steps (default = 48 = 24 hours back).
    Each entry: (from_date, from_h, from_m, to_date, to_h, to_m)
    """
    base = datetime.strptime(boundary_date, "%Y-%m-%d").replace(hour=boundary_hr, minute=0)
    windows = []
    for step in range(max_steps):
        t_start = base - timedelta(minutes=30 * step)
        t_end   = t_start + timedelta(minutes=30)
        windows.append((
            t_start.strftime("%Y-%m-%d"), t_start.hour, t_start.minute,
            t_end.strftime("%Y-%m-%d"),   t_end.hour,   t_end.minute,
        ))
    return windows


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def filename_safe(fd, fh, fm, td, th, tm):
    return f"ISS_{fd}_{fh:02d}{fm:02d}_{td}_{th:02d}{tm:02d}.xlsx"


def build_date_str(date_str: str, hour: int, minute: int, seconds: int = 0) -> str:
    """Format a date+time as IRAS expects: DD-MM-YYYY hh:mm:ss am/pm (12-hour)."""
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    am_pm = "am" if hour < 12 else "pm"
    h12   = hour % 12 or 12
    return f"{dt.day:02d}-{dt.month:02d}-{dt.year} {h12:02d}:{minute:02d}:{seconds:02d} {am_pm}"


# ─────────────────────────────────────────────
# DATE / TIME FIELD
# ─────────────────────────────────────────────

async def set_datetime_field(page, field_label: str, date_str: str, hour: int, minute: int,
                             seconds: int = 0):
    """
    Set a From Date or To Date field.
    Uses fill() for React state updates; falls back to the React native setter.
    """
    value = build_date_str(date_str, hour, minute, seconds)

    # Two date inputs share the same placeholder — From is .first, To is .last
    selector = "input[placeholder*='DD-MM-YYYY']"
    field = page.locator(selector).first if "From" in field_label else page.locator(selector).last

    await field.wait_for(state="visible", timeout=5_000)
    await field.fill(value)
    await page.wait_for_timeout(200)

    # Verify React accepted it; fall back to native setter if not
    actual = await field.input_value()
    if actual != value:
        el = await field.element_handle()
        await page.evaluate(
            "([el, val]) => {"
            "  const setter = Object.getOwnPropertyDescriptor("
            "    window.HTMLInputElement.prototype, 'value').set;"
            "  setter.call(el, val);"
            "  el.dispatchEvent(new Event('input',  {bubbles:true}));"
            "  el.dispatchEvent(new Event('change', {bubbles:true}));"
            "}",
            [el, value]
        )
        await page.wait_for_timeout(200)
        actual = await field.input_value()

    await page.keyboard.press("Tab")
    await page.wait_for_timeout(200)
    print(f"    {field_label}: {actual!r}")


# ─────────────────────────────────────────────
# NAVIGATION
# ─────────────────────────────────────────────

async def set_as_per_actual(page):
    """Select 'Actual Transaction Date and Time' in the As Per MUI Select."""
    as_per = page.locator("div[role='combobox'][aria-labelledby*='As Per']")
    await as_per.wait_for(state="visible", timeout=5_000)
    await as_per.click()
    await page.wait_for_timeout(600)
    option = page.locator("li[role='option']").filter(has_text="Actual")
    await option.first.wait_for(state="visible", timeout=5_000)
    await option.first.click()
    await page.wait_for_timeout(400)
    print("  [OK] As Per = Actual Transaction Date and Time")


async def navigate_to_iss(page, shift_date: str = SHIFT_DATES[0]):
    """
    1. Click FCC Data in the left nav
    2. Click the ... overflow tab
    3. Click Issue(ISS) in the popup
    4. If shift_date is older than 7 days, click the archive toggle
    5. Set As Per = Actual Transaction Date and Time
    """
    await page.get_by_role("button", name="FCC Data").click()
    await page.wait_for_timeout(2000)

    overflow = page.locator("button[role='tab']:has-text('...')")
    await overflow.wait_for(state="visible", timeout=10_000)
    await overflow.click()
    await page.wait_for_timeout(1000)

    iss = page.locator("li.app-tab-list:has-text('Issue(ISS)')")
    await iss.wait_for(state="visible", timeout=5_000)
    await iss.click()
    await page.wait_for_timeout(1500)

    # Click archive toggle only when date is more than 7 days in the past
    days_old = (datetime.now() - datetime.strptime(shift_date, "%Y-%m-%d")).days
    if days_old > 7:
        toggle = page.locator(".MuiFormControlLabel-root:has-text('Last 7 days Report')")
        await toggle.wait_for(state="visible", timeout=10_000)
        await toggle.click()
        await page.wait_for_timeout(600)
        print("  [OK] Archive toggle enabled (date is >7 days old)")
    else:
        print("  [--] Date within last 7 days — archive toggle not needed")

    await set_as_per_actual(page)
    print("[OK] Navigated to FCC Data > Issue(ISS)")


# ─────────────────────────────────────────────
# EXPORT ONE WINDOW
# ─────────────────────────────────────────────

async def export_window(page, output_dir: Path,
                        fd: str, fh: int, fm: int,
                        td: str, th: int, tm: int) -> bool | None:
    """
    Set the date range, click Show, and download the Excel file.

    Returns:
        True  — file downloaded (table had data)
        None  — table was empty (no records for this window, not an error)
        False — an error occurred
    """
    fname = filename_safe(fd, fh, fm, td, th, tm)
    fpath = output_dir / fname

    if fpath.exists():
        print(f"  [skip] Already exists: {fname}")
        return True

    try:
        await set_datetime_field(page, "From Date", fd, fh, fm)
        await set_datetime_field(page, "To Date",   td, th, tm)

        await page.locator("button:has-text('Show')").first.click()
        await page.wait_for_timeout(1500)

        # Wait for table to finish loading (rows present OR no-data overlay)
        try:
            await page.wait_for_selector(
                ".ag-row, .ag-overlay-no-rows-wrapper",
                timeout=TABLE_LOAD_TIMEOUT * 1000
            )
        except PlaywrightTimeout:
            pass

        await page.wait_for_timeout(500)

        row_count = await page.locator(".ag-row").count()
        if row_count == 0:
            print(f"  [empty] No data")
            return None   # not a failure

        # Download Excel
        async with page.expect_download(timeout=DOWNLOAD_TIMEOUT * 1000) as dl_info:
            await page.locator("button.export-excel-button").click()

        download = await dl_info.value
        await download.save_as(str(fpath))
        print(f"  [saved] {fname}  ({row_count} rows)")
        return True

    except PlaywrightTimeout as e:
        print(f"  [timeout] {fname}: {e}")
        return False
    except Exception as e:
        print(f"  [error] {fname}: {e}")
        return False


# ─────────────────────────────────────────────
# SESSION KEEPALIVE
# ─────────────────────────────────────────────

async def handle_session_expiry(page):
    try:
        if await page.locator("text=session").first.is_visible(timeout=1000):
            for btn in ["OK", "Continue", "Stay", "Yes"]:
                b = page.locator(f"button:has-text('{btn}')").first
                if await b.is_visible(timeout=500):
                    await b.click()
                    await page.wait_for_timeout(1000)
                    return True
    except Exception:
        pass
    return False


# ─────────────────────────────────────────────
# SHIFT TOTALIZER — NAVIGATION + DOWNLOAD
# ─────────────────────────────────────────────

async def navigate_to_shift_totalizer(page, shift_date: str):
    """Navigate to FCC Data > Shift Totalizer Record(TOT) tab (via overflow '...' menu)."""
    await page.get_by_role("button", name="FCC Data").click()
    await page.wait_for_timeout(2000)

    overflow = page.locator("button[role='tab']:has-text('...')")
    await overflow.wait_for(state="visible", timeout=10_000)
    await overflow.click()
    await page.wait_for_timeout(1000)

    st = page.locator("li.app-tab-list:has-text('Shift Totalizer Record(TOT)')")
    await st.wait_for(state="visible", timeout=5_000)
    await st.click()
    await page.wait_for_timeout(1500)

    days_old = (datetime.now() - datetime.strptime(shift_date, "%Y-%m-%d")).days
    if days_old > 7:
        toggle = page.locator(".MuiFormControlLabel-root:has-text('Last 7 days Report')")
        await toggle.wait_for(state="visible", timeout=10_000)
        await toggle.click()
        await page.wait_for_timeout(600)
        print("  [OK] Archive toggle enabled (date is >7 days old)")

    await set_as_per_actual(page)
    print("[OK] Navigated to FCC Data > Shift Totalizer Record(TOT)")


async def download_shift_totalizer(page, output_dir: Path, shift_date: str) -> "Path | None":
    """
    Download the Shift Totalizer Excel for shift_date.

    Sets date range to midnight-to-midnight on shift_date (the Shift Totalizer
    is a midnight-boundary record, not a 06:00-boundary record).

    Returns the saved file path, or None if the download fails / no data.
    """
    fname = f"ShiftTotalizer_{shift_date}.xlsx"
    fpath = output_dir / fname

    if fpath.exists():
        print(f"  [skip] Already exists: {fname}")
        return fpath

    try:
        # 12:00:00 am → 11:59:59 pm on the same calendar day
        await set_datetime_field(page, "From Date", shift_date, 0, 0, seconds=0)
        await set_datetime_field(page, "To Date",   shift_date, 23, 59, seconds=59)

        await page.locator("button:has-text('Show')").first.click()
        await page.wait_for_timeout(1500)

        try:
            await page.wait_for_selector(
                ".ag-row, .ag-overlay-no-rows-wrapper",
                timeout=TABLE_LOAD_TIMEOUT * 1000
            )
        except PlaywrightTimeout:
            pass

        await page.wait_for_timeout(500)

        row_count = await page.locator(".ag-row").count()
        if row_count == 0:
            print(f"  [empty] Shift Totalizer: no data for {shift_date}")
            return None

        async with page.expect_download(timeout=DOWNLOAD_TIMEOUT * 1000) as dl_info:
            await page.locator("button.export-excel-button").click()

        download = await dl_info.value
        await download.save_as(str(fpath))
        print(f"  [saved] {fname}  ({row_count} rows)")
        return fpath

    except PlaywrightTimeout as e:
        print(f"  [timeout] Shift Totalizer {shift_date}: {e}")
        return None
    except Exception as e:
        print(f"  [error] Shift Totalizer {shift_date}: {e}")
        return None


# ─────────────────────────────────────────────
# SHIFT TOTALIZER — PARSER
# ─────────────────────────────────────────────

def parse_shift_totalizer_nozzle(filepath: Path, nozzle_no: int) -> "dict | None":
    """
    Parse a Shift Totalizer Record(TOT) Excel file and return the open/close
    totalizer for the requested nozzle: {"open": float, "close": float}.

    File structure (confirmed from real export):
      Sheet: "Shift Totalizer Record(TOT)"
      Columns include: Nozzle No, Shift Type ('O' = open, 'C' = close), Tot Reading
      Each nozzle has two rows — one O and one C.

    Returns None if the nozzle rows are not found or the file cannot be parsed.
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"    [parse error] Could not open {filepath.name}: {e}")
        return None

    sheet = next(
        (wb[n] for n in wb.sheetnames
         if "shift" in n.lower() or "totalizer" in n.lower()),
        wb.active,
    )

    headers = None
    header_row_idx = 0
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        joined = " ".join(cells)
        if "nozzle" in joined and "shift type" in joined and "tot reading" in joined:
            headers = cells
            header_row_idx = row_idx
            break

    if headers is None:
        wb.close()
        print(f"    [parse warn] Header row not found in {filepath.name}")
        return None

    nozzle_col     = next((i for i, h in enumerate(headers) if "nozzle" in h and "no" in h),
                          next((i for i, h in enumerate(headers) if "nozzle" in h), None))
    shift_type_col = next((i for i, h in enumerate(headers) if "shift" in h and "type" in h), None)
    tot_reading_col = next((i for i, h in enumerate(headers) if "tot" in h and "reading" in h), None)

    if None in (nozzle_col, shift_type_col, tot_reading_col):
        wb.close()
        print(f"    [parse warn] Could not locate required columns in {filepath.name}")
        return None

    open_val = close_val = None

    for row in sheet.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if row[nozzle_col] is None:
            continue
        try:
            n = int(row[nozzle_col])
        except (ValueError, TypeError):
            continue
        if n != nozzle_no:
            continue

        shift_type = str(row[shift_type_col]).strip().upper() if row[shift_type_col] else ""
        try:
            reading = float(row[tot_reading_col])
        except (ValueError, TypeError):
            continue

        if shift_type == "O":
            open_val = reading
        elif shift_type == "C":
            close_val = reading

    wb.close()

    if open_val is None:
        print(f"    [parse warn] No O row found for nozzle {nozzle_no} in {filepath.name}")
        return None

    # close_val may be None if the shift hasn't ended yet (e.g. queried at 9AM same day).
    # Callers handle None close gracefully.
    return {"open": open_val, "close": close_val}


# ─────────────────────────────────────────────
# XG PRE-CHECK
# ─────────────────────────────────────────────

def xg_pre_check(st_dir: Path, shift_date: str) -> dict:
    """
    Resolve nozzle 11 (XG) before the ISS backward search runs for shift_date.
    Reads only from disk — Shift Totalizer must already be downloaded by
    download_all_shift_totalizers() before ISS navigation begins.

    We use the Shift Totalizer for (shift_date - 1 day) — the operational date
    being reconciled. Its closing value (midnight = end of calendar day) equals
    the 6AM boundary reading for shift_date, because no XG is ever sold between
    midnight and 06:00.

    Algorithm:
      1. Read the (shift_date - 1) Shift Totalizer from disk.
      2. Compute xg_movement = close - open for nozzle 11 (full calendar day movement).
      3a. If xg_movement <= 7L (pump test only):
            - Use ST close value as nozzle 11's 6AM boundary for shift_date.
            - xg_pump_test_litres = xg_movement (belongs to the op_date shift).
      3b. If xg_movement > 7L (genuine XG sale occurred during op_date):
            - Return unresolved — ISS backward search will find the correct value.
      Edge case: if file missing or parse fails, fall back to ISS search.

    Returns:
        {
            "resolved":         bool,
            "totalizer_end":    float | None,   # ST close value for nozzle 11 (when resolved)
            "pump_test_litres": float | None,   # xg_movement during op_date (when resolved)
        }
    """
    unresolved = {"resolved": False, "totalizer_end": None, "pump_test_litres": None}

    op_date_str = (datetime.strptime(shift_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
    print(f"\n[XG pre-check] {shift_date}: reading {op_date_str} Shift Totalizer from disk...")

    fpath = st_dir / f"ShiftTotalizer_{op_date_str}.xlsx"
    if not fpath.exists():
        print(f"  [XG pre-check] WARN: ShiftTotalizer_{op_date_str}.xlsx not on disk.")
        print("  [XG pre-check] Falling back to ISS boundary search for nozzle 11.")
        return unresolved

    vals = parse_shift_totalizer_nozzle(fpath, XG_NOZZLE)
    if vals is None:
        print("  [XG pre-check] WARN: Could not read nozzle 11 from Shift Totalizer.")
        print("  [XG pre-check] Falling back to ISS boundary search for nozzle 11.")
        return unresolved

    xg_movement = vals["close"] - vals["open"]
    print(f"  [XG pre-check] Nozzle 11 ({op_date_str}) — "
          f"open={vals['open']}, close={vals['close']}, movement={xg_movement:.2f}L")

    if xg_movement > XG_MOVEMENT_THRESHOLD:
        print(f"  [XG pre-check] XG movement {xg_movement:.2f}L > {XG_MOVEMENT_THRESHOLD}L threshold. "
              f"Running ISS boundary search for nozzle 11.")
        return unresolved

    # Movement ≤ 7L: pump test only, no real sale.
    # ST close (midnight of shift_date) ≈ 6AM on shift_date since no post-midnight XG sales.
    print(f"  [XG pre-check] XG exempted — movement {xg_movement:.2f}L <= "
          f"{XG_MOVEMENT_THRESHOLD}L threshold.")
    print(f"  [XG pre-check] Using {op_date_str} ST close ({vals['close']}) as nozzle 11 "
          f"6AM boundary for {shift_date}. Pump test = {xg_movement:.2f}L.")

    return {
        "resolved":         True,
        "totalizer_end":    vals["close"],
        "pump_test_litres": xg_movement,
    }


async def download_all_shift_totalizers(page, st_dir: Path, shift_dates: list):
    """
    Download Shift Totalizer files for all shift dates before ISS navigation begins.
    Called once in main() so xg_pre_check never needs to switch tabs mid-scrape.
    """
    needed = [d for d in shift_dates
              if not (st_dir / f"ShiftTotalizer_{d}.xlsx").exists()]
    if not needed:
        print("[ST] All Shift Totalizer files already on disk — skipping download.")
        return

    print(f"\n[ST] Downloading Shift Totalizer for: {needed}")
    await navigate_to_shift_totalizer(page, shift_date=needed[0])
    for shift_date in needed:
        await download_shift_totalizer(page, st_dir, shift_date)
        await asyncio.sleep(DELAY_BETWEEN)
    print("[ST] Shift Totalizer downloads complete.")


# ─────────────────────────────────────────────
# BOUNDARY MODE
# ─────────────────────────────────────────────

async def ensure_iss_archive_mode(page, shift_date: str):
    """
    Ensure the ISS 'Last 7 days Report' toggle is in the correct state for shift_date.

    - shift_date > 7 days old  → toggle must be OFF (archive mode: all history visible)
    - shift_date ≤ 7 days old  → toggle must be ON  (last-7-days mode: recent data visible)

    This matters for batch runs: processing an old date turns the toggle OFF, then a
    subsequent recent date needs it turned back ON or its dates become inaccessible.
    """
    days_old = (datetime.now() - datetime.strptime(shift_date, "%Y-%m-%d")).days
    need_archive = days_old > 7   # True  → toggle should be OFF (unchecked)
                                   # False → toggle should be ON  (checked)
    try:
        toggle   = page.locator(".MuiFormControlLabel-root:has-text('Last 7 days Report')")
        checkbox = toggle.locator("input[type='checkbox']")
        is_checked = await checkbox.is_checked(timeout=3_000)

        if need_archive and is_checked:
            await toggle.click()
            await page.wait_for_timeout(600)
            print(f"  [archive] Turned OFF 'Last 7 days' toggle — archive mode for {shift_date}")
        elif not need_archive and not is_checked:
            await toggle.click()
            await page.wait_for_timeout(600)
            print(f"  [archive] Turned ON 'Last 7 days' toggle — recent mode for {shift_date}")
        else:
            state = "archive" if not is_checked else "last-7-days"
            print(f"  [archive] Toggle already correct ({state}) for {shift_date}")
    except Exception as e:
        print(f"  [archive] Could not verify archive toggle: {e}")


async def run_boundary(page, output_dir: Path, shift_date: str) -> dict:
    """
    Boundary mode — collects the 6AM opening totalizer for all 6 nozzles.

    Step 0: XG pre-check via Shift Totalizer (resolves nozzle 11 without ISS search
            on zero/pump-test-only days; falls back to ISS if movement > 7L or no DB record).
    Step 1: ISS backward search from 05:30-06:00, stepping back in 30-min increments,
            for the 5 active nozzles only (7, 15, 16, 17, 18) — plus nozzle 11 if the
            pre-check fell back to ISS.
    Step 2: Merge XG pre-check result with ISS results into a single dict.

    Returns {nozzle_no: totalizer_end_value} for every nozzle found.
    """
    print(f"\n[boundary] {shift_date}: starting XG pre-check then ISS backward search")

    # ── Step 0: XG pre-check ────────────────────────────────────────────
    st_dir = Path(SHIFT_TOTALIZER_FOLDER)
    st_dir.mkdir(parents=True, exist_ok=True)

    xg_check = xg_pre_check(st_dir, shift_date)
    xg_resolved = xg_check["resolved"]

    # ISS search set: 5 active nozzles; add nozzle 11 back only if pre-check fell through
    remaining: set[int] = set(ACTIVE_NOZZLES)
    if not xg_resolved:
        remaining.add(XG_NOZZLE)

    # ── Step 0.5: ST pre-check for all active nozzles ────────────────────
    # If a nozzle's ST movement ≤ 7L, it was OOO or pump-test-only for the whole
    # operational day — use ST close as its 6AM boundary and skip ISS search.
    # Same logic as the XG pre-check, extended to nozzles 7, 15, 16, 17, 18.
    op_date_str = (datetime.strptime(shift_date, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
    st_fpath = st_dir / f"ShiftTotalizer_{op_date_str}.xlsx"

    st_ooo_resolved: dict[int, float] = {}  # nozzle → ST close value
    if st_fpath.exists():
        print(f"\n[ST pre-check] {shift_date}: checking active nozzles against {op_date_str} ST...")
        for nozzle in sorted(ACTIVE_NOZZLES):
            if nozzle not in remaining:
                continue
            vals = parse_shift_totalizer_nozzle(st_fpath, nozzle)
            if vals is None:
                print(f"  [ST pre-check] Nozzle {nozzle}: not in ST file — will search ISS.")
                continue
            movement = vals["close"] - vals["open"]
            print(f"  [ST pre-check] Nozzle {nozzle} ({op_date_str}) — "
                  f"open={vals['open']}, close={vals['close']}, movement={movement:.2f}L")
            if movement <= XG_MOVEMENT_THRESHOLD:
                print(f"  [ST pre-check] Nozzle {nozzle}: movement {movement:.2f}L ≤ "
                      f"{XG_MOVEMENT_THRESHOLD}L — OOO/pump-test-only. "
                      f"Using ST close {vals['close']} as 6AM boundary.")
                st_ooo_resolved[nozzle] = vals["close"]
                remaining.discard(nozzle)
            else:
                print(f"  [ST pre-check] Nozzle {nozzle}: movement {movement:.2f}L > "
                      f"{XG_MOVEMENT_THRESHOLD}L — proceeding to ISS search.")
    else:
        print(f"  [ST pre-check] WARN: {st_fpath.name} not on disk — skipping active nozzle ST check.")

    # ── Step 1: ISS backward search ─────────────────────────────────────
    found: dict[int, float] = {}

    if not remaining:
        print(f"\n[boundary] {shift_date}: all nozzles resolved by ST pre-check — skipping ISS search.")
    else:
        # Ensure archive toggle is OFF (historical data visible) — IRAS resets it on date changes
        await ensure_iss_archive_mode(page, shift_date)
        print(f"\n[boundary] {shift_date}: ISS backward search for nozzles {sorted(remaining)}")
        wins = boundary_windows(shift_date, SHIFT_START_HR)

        for (fd, fh, fm, td, th, tm) in wins[1:]:   # skip step-0 (06:00-06:30 = new shift)
            if not remaining:
                break

            print(f"  [{shift_date}] checking {fd} {fh:02d}:{fm:02d}-{td} {th:02d}:{tm:02d}  "
                  f"(still need nozzles: {sorted(remaining)})")

            await handle_session_expiry(page)
            result = await export_window(page, output_dir, fd, fh, fm, td, th, tm)

            if result is True:
                fpath = output_dir / filename_safe(fd, fh, fm, td, th, tm)
                window_data = parse_totalizer_ends(fpath)
                for nozzle, tot_end in window_data.items():
                    if nozzle in remaining:
                        found[nozzle] = tot_end
                        remaining.discard(nozzle)
                        print(f"    [+] Nozzle {nozzle}: Totalizer End = {tot_end}")

            await asyncio.sleep(DELAY_BETWEEN)

        if remaining:
            print(f"  [WARN] No ISS data found for nozzles: {sorted(remaining)} — attempting carry-forward from DB")
            carry = lookup_carry_forward(remaining, shift_date)
            found.update(carry)
            remaining -= set(carry.keys())
        if remaining:
            print(f"  [WARN] Could not resolve nozzles {sorted(remaining)} — they will be missing from DB")

    # ── Step 2: merge ST pre-check results (XG + OOO active nozzles) ───
    if xg_resolved:
        found[XG_NOZZLE] = xg_check["totalizer_end"]
        print(f"    [+] Nozzle {XG_NOZZLE} (XG ST carry-forward): "
              f"Totalizer End = {xg_check['totalizer_end']}")

    for nozzle, tot_end in st_ooo_resolved.items():
        found[nozzle] = tot_end
        print(f"    [+] Nozzle {nozzle} (OOO ST carry-forward): Totalizer End = {tot_end}")

    print(f"\n  [boundary result] {shift_date}: {found}")
    return found, xg_check   # return xg_check so save_totalizers_to_db can use the override


# ─────────────────────────────────────────────
# FULL SHIFT MODE
# ─────────────────────────────────────────────

async def run_full_shift(page, output_dir: Path, shift_date: str):
    """Export all 48 half-hour windows of the 06:00-to-06:00 shift."""
    windows = shift_windows(shift_date, SHIFT_START_HR)
    total = len(windows)
    success = skipped = failed = 0

    print(f"\n[full shift] {shift_date}  06:00 -> next day 06:00  ({total} windows)")

    for (fd, fh, fm, td, th, tm) in windows:
        print(f"  {fd} {fh:02d}:{fm:02d} - {td} {th:02d}:{tm:02d}", end="  ")
        await handle_session_expiry(page)
        result = await export_window(page, output_dir, fd, fh, fm, td, th, tm)
        if result is True:
            success += 1
        elif result is None:
            skipped += 1
        else:
            failed += 1
        await asyncio.sleep(DELAY_BETWEEN)

    print(f"\n  Shift {shift_date}: {success} downloaded, {skipped} empty, {failed} errors")
    return success, skipped, failed


# ─────────────────────────────────────────────
# PUMP TEST SCAN MODE
# ─────────────────────────────────────────────

def pump_test_windows(shift_date: str):
    """Return the 7 half-hour windows from 08:00 to 11:30 for a given date."""
    base = datetime.strptime(shift_date, "%Y-%m-%d").replace(hour=8, minute=0)
    windows = []
    t = base
    while t.hour < 11 or (t.hour == 11 and t.minute < 30):
        t_end = t + timedelta(minutes=30)
        windows.append((
            t.strftime("%Y-%m-%d"),   t.hour,   t.minute,
            t_end.strftime("%Y-%m-%d"), t_end.hour, t_end.minute,
        ))
        t = t_end
    return windows


async def run_pump_test_scan(page, output_dir: Path, shift_dates: list):
    """
    Download the 08:00-11:30 ISS windows for each date and print a pump test summary.
    Only 7 windows per date — no wasted requests outside that range.
    """
    print(f"\n[pump test scan] dates: {shift_dates}  window: 08:00–11:30")

    for shift_date in shift_dates:
        wins = pump_test_windows(shift_date)
        print(f"\n  {shift_date}: downloading {len(wins)} windows...")
        for (fd, fh, fm, td, th, tm) in wins:
            await handle_session_expiry(page)
            await export_window(page, output_dir, fd, fh, fm, td, th, tm)
            await asyncio.sleep(DELAY_BETWEEN)

    # Scan all downloaded files and print summary
    print("\n" + "=" * 60)
    print(f"  PUMP TEST SUMMARY  ({shift_dates[0]} → {shift_dates[-1]})")
    print("=" * 60)
    print(f"  {'Date':<12} {'Nozzle':>7}  {'Pump Test (L)':>13}")
    print(f"  {'-'*12} {'-'*7}  {'-'*13}")

    for shift_date in shift_dates:
        op_date = datetime.strptime(shift_date, "%Y-%m-%d").date()
        results = scan_pump_tests(op_date, output_dir)
        if results:
            for nozzle in sorted(results):
                print(f"  {shift_date:<12} {nozzle:>7}  {results[nozzle]:>13.2f}")
        else:
            print(f"  {shift_date:<12}    —     no pump test rows found")

    print("=" * 60)


# ─────────────────────────────────────────────
# LOGIN WAIT
# ─────────────────────────────────────────────

async def auto_fill_login(page):
    """
    1. Select 'Dealer' from the role dropdown.
    2. Fill Username and Password.
    Leaves CAPTCHA for manual entry.
    """
    try:
        await page.wait_for_timeout(1000)

        # Try native <select> first
        sel = page.locator("select").first
        if await sel.count() > 0 and await sel.is_visible(timeout=2000):
            await sel.select_option(label="Dealer")
            print("[OK] Selected Dealer via <select>")
        else:
            # MUI-style dropdown — click to open, then pick option
            dropdown = page.locator("div[role='combobox'], .MuiSelect-select").first
            await dropdown.wait_for(state="visible", timeout=5000)
            await dropdown.click()
            await page.wait_for_timeout(600)
            dealer_opt = page.locator("li[role='option']:has-text('Dealer'), option:has-text('Dealer')").first
            await dealer_opt.wait_for(state="visible", timeout=5000)
            await dealer_opt.click()
            print("[OK] Selected Dealer via MUI dropdown")

        await page.wait_for_timeout(800)

        # Fill Username (206858) — appears after Dealer is selected
        for sel in ["input[name='username']", "input[name='userId']",
                    "input[placeholder*='Username']", "input[placeholder*='username']",
                    "input[placeholder*='User']"]:
            un = page.locator(sel).first
            if await un.count() > 0 and await un.is_visible(timeout=1000):
                await un.fill(USERNAME)
                print(f"[OK] Username filled ({USERNAME})")
                break

        # Fill Password
        pw = page.locator("input[type='password']").first
        if await pw.count() > 0 and await pw.is_visible(timeout=2000):
            await pw.fill(PASSWORD)
            print("[OK] Password filled")

    except Exception as e:
        print(f"[--] Auto-fill failed: {e}")
        print("     Please fill in Dealer / Username / Password manually.")


async def wait_for_login(page):
    await auto_fill_login(page)

    print()
    print("=" * 55)
    print("  ACTION REQUIRED")
    print("=" * 55)
    print("  Fill in the CAPTCHA and click Login")
    print("  (script waits up to 5 minutes)")
    print("=" * 55)
    print()
    await page.wait_for_function(
        "() => !window.location.href.includes('/login')",
        timeout=300_000
    )
    print("[OK] Login detected")
    await page.wait_for_timeout(2000)


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────

async def main():
    output_dir = Path(OUTPUT_FOLDER)
    output_dir.mkdir(parents=True, exist_ok=True)

    print()
    print("IRAS ISS Exporter")
    print(f"Mode        : {RUN_MODE}")
    print(f"Shift dates : {SHIFT_DATES}")
    print(f"Output      : {OUTPUT_FOLDER}")
    print()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, args=["--start-maximized"])
        _iras_proxy = iras_proxy_cfg()
        _ctx_kw: dict = {
            "accept_downloads": True,
            "viewport": {"width": 1400, "height": 900},
        }
        if _iras_proxy is not None:
            _ctx_kw["proxy"] = _iras_proxy
        # Context/page setup — exceptions suppressed to avoid leaking proxy config.
        try:
            context = await browser.new_context(**_ctx_kw)
            page = await context.new_page()
        except Exception as _setup_exc:
            print(f"  [IRAS] Browser/context setup failed: {safe_exc_name(_setup_exc)}")
            await browser.close()
            return

        # Login
        print(f"  [IRAS] proxy : {'yes' if IRAS_PROXY_ENABLED else 'no'}")
        # Initial navigation — raw error message suppressed (may contain proxy host/port).
        try:
            await page.goto(IRAS_URL, wait_until="networkidle")
        except PlaywrightTimeout:
            print(f"  [IRAS] Navigation timeout (networkidle) — continuing")
        except Exception as _nav_exc:
            print(f"  [IRAS] Initial navigation failed: {safe_exc_name(_nav_exc)}")
            await browser.close()
            return
        await page.wait_for_timeout(1500)

        await wait_for_login(page)

        # Download Shift Totalizers for every op_date in the batch.
        # Each boundary date's op_date = boundary_date - 1 day.
        # download_all_shift_totalizers skips files already on disk.
        st_dir = Path(SHIFT_TOTALIZER_FOLDER)
        st_dir.mkdir(parents=True, exist_ok=True)
        op_dates = [
            (datetime.strptime(d, "%Y-%m-%d") - timedelta(days=1)).strftime("%Y-%m-%d")
            for d in SHIFT_DATES
        ]
        await download_all_shift_totalizers(page, st_dir, op_dates)

        # Navigate to ISS for the boundary search
        await navigate_to_iss(page, shift_date=SHIFT_DATES[0])

        # Run selected mode
        if RUN_MODE == "pump_test_scan":
            await run_pump_test_scan(page, output_dir, SHIFT_DATES)
        else:
            for shift_date in SHIFT_DATES:
                if RUN_MODE == "boundary":
                    totalizers, xg_check = await run_boundary(page, output_dir, shift_date)
                    if totalizers:
                        print(f"\n  --> 6AM opening totalizers for {shift_date}:")
                        for nozzle in sorted(totalizers):
                            print(f"       Nozzle {nozzle}: {totalizers[nozzle]}")
                        save_totalizers_to_db(shift_date, totalizers, xg_check)
                    else:
                        print(f"\n  --> [WARN] No totalizer data found for {shift_date}")
                elif RUN_MODE == "full":
                    await run_full_shift(page, output_dir, shift_date)

        print("\n[DONE]")
        await browser.close()


def scan_pump_tests(op_date, output_folder: Path) -> dict:
    """
    Scan all downloaded ISS files for the operational day and return
    {nozzle_no: total_pump_test_litres} from Pump Test (105) transactions.

    Operational day window: op_date 06:00 → op_date+1 06:00.
    Only files already on disk are scanned — no new downloads.
    """
    from datetime import timedelta, datetime as _dt

    op_start = _dt.combine(op_date, _dt.min.time().replace(hour=6, minute=0))
    op_end   = op_start + timedelta(hours=24)

    pump_tests: dict[int, float] = {}

    for fpath in sorted(output_folder.glob("ISS_*.xlsx")):
        # Filename format: ISS_{from_date}_{from_hhmm}_{to_date}_{to_hhmm}.xlsx
        # e.g. ISS_2026-04-16_0800_2026-04-16_0830.xlsx
        parts = fpath.stem.split("_")
        try:
            from_dt = _dt.strptime(f"{parts[1]} {parts[2]}", "%Y-%m-%d %H%M")
        except (IndexError, ValueError):
            continue

        # Only include windows that start within the operational day
        if not (op_start <= from_dt < op_end):
            continue

        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception:
            continue

        sheet = next(
            (wb[n] for n in wb.sheetnames if "iss" in n.lower() or "issue" in n.lower()),
            wb.active,
        )

        # Find header row
        headers = None
        hdr_idx = 0
        for idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
            cells = [str(c).strip().lower() if c is not None else "" for c in row]
            if "nozzle" in " ".join(cells) and "totalizer" in " ".join(cells):
                headers = cells
                hdr_idx = idx
                break

        if not headers:
            wb.close()
            continue

        txn_col    = next((i for i, h in enumerate(headers) if "iss transaction type" in h or
                           ("transaction" in h and "type" in h and "iss" in " ".join(headers[:i+1]))), None)
        nozzle_col = next((i for i, h in enumerate(headers) if "nozzle" in h and "no" in h),
                          next((i for i, h in enumerate(headers) if "nozzle" in h), None))
        qty_col    = next((i for i, h in enumerate(headers) if h == "quantity"), None)

        if txn_col is None or nozzle_col is None or qty_col is None:
            wb.close()
            continue

        for row in sheet.iter_rows(min_row=hdr_idx + 1, values_only=True):
            if not any(row):
                continue
            txn_type = str(row[txn_col]).strip() if row[txn_col] is not None else ""
            if "pump test" not in txn_type.lower():
                continue
            try:
                nozzle = int(row[nozzle_col])
                qty    = float(row[qty_col]) if row[qty_col] is not None else 0.0
            except (TypeError, ValueError):
                continue
            pump_tests[nozzle] = pump_tests.get(nozzle, 0.0) + qty

        wb.close()

    return pump_tests


def lookup_carry_forward(nozzles: set, shift_date: str) -> dict:
    """
    For nozzles that had zero transactions during the backward search window,
    carry forward the most recent NozzleTotalizer value from the DB.
    Valid because a totalizer only changes when fuel is dispensed — if no
    transactions were found in 48 windows, the reading hasn't changed.
    """
    if not nozzles:
        return {}
    try:
        from pumpvision import create_app
        from pumpvision.models import db, NozzleTotalizer
        app = create_app()
        result = {}
        with app.app_context():
            op_date = datetime.strptime(shift_date, "%Y-%m-%d").date()
            for nozzle in nozzles:
                prev = (
                    db.session.query(NozzleTotalizer)
                    .filter(
                        NozzleTotalizer.nozzle_no == nozzle,
                        NozzleTotalizer.operational_date < op_date,
                    )
                    .order_by(NozzleTotalizer.operational_date.desc())
                    .first()
                )
                if prev:
                    result[nozzle] = prev.totalizer_end
                    print(f"    [carry-forward] Nozzle {nozzle}: {prev.totalizer_end} "
                          f"from {prev.operational_date} (zero activity in search window)")
                else:
                    print(f"    [carry-forward] Nozzle {nozzle}: no prior DB entry — cannot resolve")
        return result
    except Exception as e:
        print(f"  [carry-forward error] {e}")
        return {}


def save_totalizers_to_db(shift_date: str, totalizers: dict, xg_check: dict = None):
    """
    Write boundary totalizer results to the pumpvision DB.
    Creates or updates NozzleTotalizer rows for the given operational date.

    xg_check: the dict returned by xg_pre_check(). When XG was resolved via
    carry-forward (resolved=True), its pump_test_litres come from the Shift
    Totalizer movement — not from the ISS scan — and are applied here as an
    override so we don't incorrectly store 0 for nozzle 11's pump test.
    """
    if not totalizers:
        print("  [db] Nothing to save.")
        return

    try:
        from pumpvision import create_app
        from pumpvision.models import db, NozzleTotalizer

        app = create_app()
        with app.app_context():
            op_date = datetime.strptime(shift_date, "%Y-%m-%d").date()

            # Pump tests for the 5 active nozzles: scan ISS files for the previous
            # operational day window (the shift that ended at op_date 06:00).
            prev_op_date = op_date - timedelta(days=1)
            pump_tests = scan_pump_tests(prev_op_date, Path(OUTPUT_FOLDER))
            if pump_tests:
                print(f"  [db] Pump test litres found: { {n: v for n, v in pump_tests.items()} }")
            else:
                print(f"  [db] No pump test rows found for {shift_date}")

            saved = 0
            for nozzle_no, totalizer_end in totalizers.items():
                existing = db.session.query(NozzleTotalizer).filter_by(
                    operational_date=op_date, nozzle_no=nozzle_no
                ).first()

                # Determine pump test litres for this nozzle.
                # For XG resolved via carry-forward, use the Shift Totalizer movement
                # (the pump test wasn't in any ISS file we downloaded today).
                if (nozzle_no == XG_NOZZLE
                        and xg_check is not None
                        and xg_check.get("resolved")
                        and xg_check.get("pump_test_litres") is not None):
                    pt_litres = xg_check["pump_test_litres"]
                else:
                    pt_litres = pump_tests.get(nozzle_no, 0.0)

                if existing:
                    existing.totalizer_end    = float(totalizer_end)
                    existing.pump_test_litres = pt_litres
                    existing.scraped_at       = datetime.utcnow()
                else:
                    db.session.add(NozzleTotalizer(
                        operational_date=op_date,
                        nozzle_no=nozzle_no,
                        totalizer_end=float(totalizer_end),
                        pump_test_litres=pt_litres,
                    ))
                saved += 1
            db.session.commit()
            print(f"  [db] Saved {saved} nozzle totalizer readings for {shift_date}")
    except Exception as e:
        print(f"  [db] ERROR writing to DB: {e}")


if __name__ == "__main__":
    asyncio.run(main())
