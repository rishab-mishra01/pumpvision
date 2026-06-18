"""
IRAS Price (PRM) Exporter
==========================
Exports the FCC Data > Price(PRM) tab from the IRAS portal for a given date range.

Unlike the ISS exporter there is no 30-minute window limit — the full date range
is exported in a single Excel file.  The script also parses the downloaded file
and prints a clean summary of RSP per product per operational day.

USAGE:
    python -X utf8 iras_price_exporter.py
"""

import asyncio
import io
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout
from iras_proxy import iras_proxy_cfg, IRAS_PROXY_ENABLED, safe_exc_name

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")


# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────

IRAS_URL  = "https://iras.iocliras.in/login"
USERNAME  = "206858"
PASSWORD  = "Shree@26"

OUTPUT_FOLDER = r"C:\IRAS_Data\Price"

# Date range to export — inclusive.
# For daily reconciliation, run once a week or once a month.
FROM_DATE = "2026-04-01"   # YYYY-MM-DD
TO_DATE   = "2026-04-16"   # YYYY-MM-DD

TABLE_LOAD_TIMEOUT = 30   # seconds to wait for table after clicking Show
DOWNLOAD_TIMEOUT   = 30   # seconds to wait for Excel download


# ─────────────────────────────────────────────
# EXCEL PARSER
# ─────────────────────────────────────────────

def parse_price_file(filepath: Path) -> list[dict]:
    """
    Parse a Price(PRM) Excel export and return a list of price records.

    Each record:
        {
            "product":        str,       # HS / MS / X2 / XG
            "rate_per_litre": float,
            "effective_from": datetime,  # 06:00:00 on the operational day
            "effective_to":   datetime,  # 05:59:59 the following day
        }

    Rows with Record Type != 'RSP' or unknown products are skipped.
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as e:
        print(f"  [parse error] Could not open {filepath.name}: {e}")
        return []

    # Find the Price(PRM) sheet
    sheet_name = next(
        (n for n in wb.sheetnames if "price" in n.lower() or "prm" in n.lower()),
        wb.sheetnames[0],
    )
    ws = wb[sheet_name]

    # Locate the header row by looking for 'product' and 'rate'
    headers = None
    header_row_idx = 0
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        joined = " ".join(cells)
        if "product" in joined and "rate" in joined:
            headers = cells
            header_row_idx = row_idx
            break

    if headers is None:
        wb.close()
        print(f"  [parse warn] Header row not found in {filepath.name}")
        return []

    def col(keyword1, keyword2=None):
        """Find column index containing keyword1 (and optionally keyword2)."""
        for i, h in enumerate(headers):
            if keyword1 in h:
                if keyword2 is None or keyword2 in h:
                    return i
        return None

    record_type_col  = col("record")
    product_col      = col("product", "code")
    rate_col         = col("rate")
    date_from_col    = col("effective", "from") if col("effective date from") is None else col("effective date from")
    time_from_col    = col("effective time from")
    date_to_col      = col("effective date to") if col("effective", "to") is not None else col("effective", "to")
    time_to_col      = col("effective time upto") if col("time upto") is not None else col("time", "upto")

    # Fallback column resolution using positional search
    if date_from_col is None:
        date_from_col = next((i for i, h in enumerate(headers) if "date" in h and "from" in h), None)
    if time_from_col is None:
        time_from_col = next((i for i, h in enumerate(headers) if "time" in h and "from" in h), None)
    if date_to_col is None:
        date_to_col = next((i for i, h in enumerate(headers) if "date" in h and "to" in h), None)
    if time_to_col is None:
        time_to_col = next((i for i, h in enumerate(headers) if "time" in h and ("to" in h or "upto" in h)), None)

    required = {
        "product_col": product_col,
        "rate_col": rate_col,
        "date_from_col": date_from_col,
        "time_from_col": time_from_col,
    }
    missing = [k for k, v in required.items() if v is None]
    if missing:
        wb.close()
        print(f"  [parse warn] Could not locate columns: {missing} in {filepath.name}")
        return []

    KNOWN_PRODUCTS = {"HS", "MS", "X2", "XG"}
    records = []

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # Skip if record type column exists and it's not RSP
        if record_type_col is not None:
            rec_type = str(row[record_type_col]).strip().upper() if row[record_type_col] else ""
            if rec_type and rec_type != "RSP":
                continue

        product = str(row[product_col]).strip().upper() if row[product_col] else ""
        if product not in KNOWN_PRODUCTS:
            continue

        try:
            rate = float(row[rate_col])
        except (TypeError, ValueError):
            continue

        # Parse effective_from datetime
        try:
            date_from_str = str(row[date_from_col]).strip()   # DD-MM-YYYY
            time_from_str = str(row[time_from_col]).strip()   # HH:MM:SS
            effective_from = datetime.strptime(
                f"{date_from_str} {time_from_str}", "%d-%m-%Y %H:%M:%S"
            )
        except (TypeError, ValueError):
            continue

        # Parse effective_to datetime (optional — may be absent)
        effective_to = None
        if date_to_col is not None and time_to_col is not None:
            try:
                date_to_str = str(row[date_to_col]).strip()
                time_to_str = str(row[time_to_col]).strip()
                effective_to = datetime.strptime(
                    f"{date_to_str} {time_to_str}", "%d-%m-%Y %H:%M:%S"
                )
            except (TypeError, ValueError):
                pass

        records.append({
            "product":        product,
            "rate_per_litre": rate,
            "effective_from": effective_from,
            "effective_to":   effective_to,
        })

    wb.close()
    return records


def print_price_summary(records: list[dict]):
    """Print a human-readable table of the parsed price records."""
    if not records:
        print("  [no records]")
        return

    # Sort by effective_from then product
    records_sorted = sorted(records, key=lambda r: (r["effective_from"], r["product"]))

    print()
    print(f"  {'Product':<10} {'Rate (₹/L)':>12}  {'Effective From':<22}  {'Effective To'}")
    print(f"  {'-'*10} {'-'*12}  {'-'*22}  {'-'*22}")
    for r in records_sorted:
        to_str = r["effective_to"].strftime("%d-%b-%Y %H:%M:%S") if r["effective_to"] else "—"
        print(
            f"  {r['product']:<10} {r['rate_per_litre']:>12.2f}  "
            f"{r['effective_from'].strftime('%d-%b-%Y %H:%M:%S'):<22}  {to_str}"
        )
    print()


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

def build_date_str(date_str: str, hour: int, minute: int) -> str:
    """Format a date+time as IRAS expects: DD-MM-YYYY hh:mm:ss am/pm (12-hour)."""
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    am_pm = "am" if hour < 12 else "pm"
    h12   = hour % 12 or 12
    return f"{dt.day:02d}-{dt.month:02d}-{dt.year} {h12:02d}:{minute:02d}:00 {am_pm}"


async def set_datetime_field(page, field_label: str, date_str: str, hour: int, minute: int):
    """Set a From Date or To Date field (same React input pattern as ISS scraper)."""
    value = build_date_str(date_str, hour, minute)

    selector = "input[placeholder*='DD-MM-YYYY']"
    field = page.locator(selector).first if "From" in field_label else page.locator(selector).last

    await field.wait_for(state="visible", timeout=5_000)
    await field.fill(value)
    await page.wait_for_timeout(200)

    actual = await field.input_value()
    if actual != value:
        el = await field.element_handle()
        await page.evaluate(
            "([el, val]) => {"
            "  const setter = Object.getOwnPropertyDescriptor("
            "    window.HTMLInputElement.prototype, 'value').set;"
            "  setter.call(el, val);"
            "  el.dispatchEvent(new Event('input',  {bubbles:true}));"
            "  el.dispatchEvent(new Event('change', {bubbles:true}));"
            "}",
            [el, value]
        )
        await page.wait_for_timeout(200)
        actual = await field.input_value()

    await page.keyboard.press("Tab")
    await page.wait_for_timeout(200)
    print(f"    {field_label}: {actual!r}")


# ─────────────────────────────────────────────
# NAVIGATION
# ─────────────────────────────────────────────

async def navigate_to_price(page):
    """
    1. Click FCC Data in the left nav
    2. Wait for the tab bar to actually render (first tab visible = panel is ready)
    3. Click Price(PRM) directly if visible, otherwise use the '...' overflow menu.
    No archive toggle and no As Per dropdown on this tab.
    """
    # Wait for the left nav to be ready before the first click. After login the
    # SPA dashboard may not have mounted the nav buttons yet, so we wait up to
    # 20s for the FCC Data button to become visible rather than clicking blindly.
    try:
        await page.get_by_role("button", name="FCC Data").wait_for(
            state="visible", timeout=20_000
        )
    except Exception:
        print("  [nav] WARNING: FCC Data button not visible after 20s — proceeding anyway")

    for attempt in range(1, 4):
        if attempt > 1:
            # On retry the SPA is in a partial state — navigate back to the
            # dashboard root to reset before re-clicking FCC Data.
            try:
                _base = IRAS_URL.replace("/login", "")
                await page.goto(_base, wait_until="networkidle", timeout=20_000)
            except Exception:
                pass
            await page.wait_for_timeout(2000)
            # Wait for nav to be ready again after reset.
            try:
                await page.get_by_role("button", name="FCC Data").wait_for(
                    state="visible", timeout=15_000
                )
            except Exception:
                pass

        await page.get_by_role("button", name="FCC Data").click()

        # Wait for the tab bar to render rather than using a fixed sleep.
        # Any button[role='tab'] appearing means the FCC panel is loaded.
        try:
            await page.locator("button[role='tab']").first.wait_for(
                state="visible", timeout=20_000
            )
        except Exception:
            print(f"  [nav] FCC Data panel did not render (attempt {attempt}/3) — retrying")
            await page.wait_for_timeout(2000)
            continue

        await page.wait_for_timeout(500)

        # Try Price(PRM) as a direct tab first
        price_tab = page.locator("button[role='tab']:has-text('Price(PRM)')")
        try:
            await price_tab.wait_for(state="visible", timeout=3_000)
            await price_tab.click()
            break
        except Exception:
            pass

        # Fall back to the '...' overflow menu
        overflow = page.locator("button[role='tab']:has-text('...')")
        try:
            await overflow.wait_for(state="visible", timeout=8_000)
            await overflow.click()
            await page.wait_for_timeout(800)
            prm = page.locator("li.app-tab-list:has-text('Price(PRM)')")
            await prm.wait_for(state="visible", timeout=5_000)
            await prm.click()
            break
        except Exception as e:
            print(f"  [nav] Price(PRM) not found via overflow (attempt {attempt}/3): {e}")
            if attempt == 3:
                raise
            await page.wait_for_timeout(2000)

    await page.wait_for_timeout(1500)
    print("[OK] Navigated to FCC Data > Price(PRM)")


# ─────────────────────────────────────────────
# EXPORT
# ─────────────────────────────────────────────

async def export_price_range(page, output_dir: Path,
                             from_date: str, to_date: str) -> Path | None:
    """
    Set the date range, click Show, and download the Price Excel.

    Returns the saved file path on success, None on failure.
    """
    from_dt = datetime.strptime(from_date, "%Y-%m-%d")
    to_dt   = datetime.strptime(to_date,   "%Y-%m-%d")

    fname = f"Price_{from_dt.strftime('%Y%m%d')}_{to_dt.strftime('%Y%m%d')}.xlsx"
    fpath = output_dir / fname

    if fpath.exists():
        print(f"  [skip] Already exists: {fname}")
        return fpath

    try:
        # From Date: start of the first operational day (06:00)
        await set_datetime_field(page, "From Date", from_date, 6, 0)
        # To Date: end of the last operational day (05:59 next calendar day)
        # Using 23:59 on to_date is sufficient — IRAS will include all prices up to end of day
        await set_datetime_field(page, "To Date", to_date, 23, 59)

        await page.locator("button:has-text('Show')").first.click()
        await page.wait_for_timeout(1500)

        try:
            await page.wait_for_selector(
                ".ag-row, .ag-overlay-no-rows-wrapper",
                timeout=TABLE_LOAD_TIMEOUT * 1000
            )
        except PlaywrightTimeout:
            pass

        await page.wait_for_timeout(500)

        row_count = await page.locator(".ag-row").count()
        if row_count == 0:
            print(f"  [empty] No price data returned for {from_date} → {to_date}")
            return None

        async with page.expect_download(timeout=DOWNLOAD_TIMEOUT * 1000) as dl_info:
            await page.locator("button.export-excel-button").click()

        download = await dl_info.value
        await download.save_as(str(fpath))
        print(f"  [saved] {fname}  ({row_count} rows)")
        return fpath

    except PlaywrightTimeout as e:
        print(f"  [timeout] {fname}: {e}")
        return None
    except Exception as e:
        print(f"  [error] {fname}: {e}")
        return None


# ─────────────────────────────────────────────
# LOGIN WAIT
# ─────────────────────────────────────────────

async def wait_for_login(page):
    print()
    print("=" * 55)
    print("  MANUAL LOGIN REQUIRED")
    print("=" * 55)
    print("  1. Select Dealer from the dropdown")
    print("  2. Confirm username / password")
    print("  3. Solve the CAPTCHA")
    print("  4. Click Login")
    print("  (script waits up to 5 minutes)")
    print("=" * 55)
    print()
    await page.wait_for_function(
        "() => !window.location.href.includes('/login')",
        timeout=300_000
    )
    print("[OK] Login detected")
    await page.wait_for_timeout(2000)


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────

async def main():
    output_dir = Path(OUTPUT_FOLDER)
    output_dir.mkdir(parents=True, exist_ok=True)

    print()
    print("IRAS Price (PRM) Exporter")
    print(f"Date range  : {FROM_DATE} → {TO_DATE}")
    print(f"Output      : {OUTPUT_FOLDER}")
    print()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False, args=["--start-maximized"])
        _iras_proxy = iras_proxy_cfg()
        _ctx_kw: dict = {
            "accept_downloads": True,
            "viewport": {"width": 1400, "height": 900},
        }
        if _iras_proxy is not None:
            _ctx_kw["proxy"] = _iras_proxy
        # Context/page setup — exceptions suppressed to avoid leaking proxy config.
        try:
            context = await browser.new_context(**_ctx_kw)
            page = await context.new_page()
        except Exception as _setup_exc:
            print(f"  [IRAS] Browser/context setup failed: {safe_exc_name(_setup_exc)}")
            await browser.close()
            return
        print(f"  [IRAS] proxy : {'yes' if IRAS_PROXY_ENABLED else 'no'}")
        # Initial navigation — raw error message suppressed (may contain proxy host/port).
        try:
            await page.goto(IRAS_URL, wait_until="networkidle")
        except PlaywrightTimeout:
            print(f"  [IRAS] Navigation timeout (networkidle) — continuing")
        except Exception as _nav_exc:
            print(f"  [IRAS] Initial navigation failed: {safe_exc_name(_nav_exc)}")
            await browser.close()
            return
        await page.wait_for_timeout(1500)

        # Pre-fill credentials
        try:
            for sel in ["input[name='username']", "input[name='userId']",
                        "input[placeholder*='User']", "input[placeholder*='user']",
                        "input[type='text']"]:
                un = page.locator(sel).first
                if await un.count() > 0 and await un.is_visible(timeout=1000):
                    await un.fill(USERNAME)
                    break
            await page.locator("input[type='password']").fill(PASSWORD)
            print("[OK] Credentials pre-filled")
        except Exception as e:
            print(f"[--] Could not pre-fill credentials: {e}")

        await wait_for_login(page)
        await navigate_to_price(page)

        fpath = await export_price_range(page, output_dir, FROM_DATE, TO_DATE)

        if fpath:
            print("\nParsing downloaded file...")
            records = parse_price_file(fpath)
            print(f"  {len(records)} price records parsed")
            print_price_summary(records)
        else:
            print("\n[WARN] No file downloaded — nothing to parse")

        print("[DONE]")
        await browser.close()


if __name__ == "__main__":
    asyncio.run(main())
