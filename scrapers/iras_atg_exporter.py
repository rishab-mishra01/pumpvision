"""
IRAS ATG Stock Exporter
========================
Scrapes the FCC Data > Stock tab from the IRAS portal and writes ATG tank
level snapshots to the `tank_readings` database table.

The Stock tab shows the current ATG reading per tank, refreshed every ~30 min
by the GVR MAG PLUS probes. Each run captures one point-in-time snapshot.

Tank map (static, outlet-specific):
  Tank 1 → HS   20,000 L
  Tank 2 → MS   20,000 L
  Tank 3 → X2   10,000 L
  Tank 4 → XG   20,000 L  — probe unreliable, is_reliable=False

Usage (standalone, manual login):
    python -X utf8 scrapers/iras_atg_exporter.py

Integrated use:
    Called from daily_scrape.py as Job 5 with an active authenticated page.
    Entry point: run_atg(page) — returns list of saved TankReading rows.
"""

import asyncio
import base64
import io
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from playwright.async_api import async_playwright, TimeoutError as PlaywrightTimeout
from iras_proxy import iras_proxy_cfg, IRAS_PROXY_ENABLED, safe_exc_name

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────

IRAS_URL = "https://iras.iocliras.in/login"

# product code → tank_id, capacity, reliability
PRODUCT_TANK_MAP = {
    "HS": {"tank_id": 1, "capacity_litres": 20000.0, "is_reliable": True},
    "MS": {"tank_id": 2, "capacity_litres": 20000.0, "is_reliable": True},
    "X2": {"tank_id": 3, "capacity_litres": 10000.0, "is_reliable": True},
    "XG": {"tank_id": 4, "capacity_litres": 20000.0, "is_reliable": False},
}

TABLE_LOAD_TIMEOUT = 30   # seconds
DOWNLOAD_TIMEOUT   = 30   # seconds


# ─────────────────────────────────────────────
# NAVIGATION
# ─────────────────────────────────────────────

async def navigate_to_stock(page):
    """Navigate to FCC Data > Stock tab."""
    await page.get_by_role("button", name="FCC Data").click()
    await page.wait_for_timeout(2000)

    # Try direct tab first (visible without overflow)
    stock_tab = page.locator("button[role='tab']:has-text('Stock')")
    try:
        await stock_tab.wait_for(state="visible", timeout=4_000)
        await stock_tab.click()
    except Exception:
        # Stock is in the overflow '...' menu — same pattern as Shift Totalizer
        overflow = page.locator("button[role='tab']:has-text('...')")
        await overflow.wait_for(state="visible", timeout=10_000)
        await overflow.click()
        await page.wait_for_timeout(800)
        stock_item = page.locator("li.app-tab-list:has-text('Stock')")
        await stock_item.wait_for(state="visible", timeout=5_000)
        await stock_item.click()

    await page.wait_for_timeout(1500)
    print("[OK] Navigated to FCC Data > Stock")


# ─────────────────────────────────────────────
# AG-GRID READER
# ─────────────────────────────────────────────

async def _read_ag_grid(page) -> list[dict]:
    """
    Read all visible ag-Grid rows into a list of {header: value} dicts.
    Returns empty list if headers cannot be found.
    """
    try:
        header_cells = page.locator(".ag-header-cell-text")
        count = await header_cells.count()
        headers = []
        for i in range(count):
            text = (await header_cells.nth(i).text_content() or "").strip().lower()
            headers.append(text)

        if not headers:
            return []

        rows = []
        row_els = page.locator(".ag-row")
        row_count = await row_els.count()

        for r in range(row_count):
            cells = row_els.nth(r).locator(".ag-cell")
            cell_count = await cells.count()
            row_data = {}
            for c in range(min(cell_count, len(headers))):
                val = (await cells.nth(c).text_content() or "").strip()
                row_data[headers[c]] = val
            if any(row_data.values()):
                rows.append(row_data)

        return rows

    except Exception as e:
        print(f"  [ATG] ag-Grid read error: {e}")
        return []


# ─────────────────────────────────────────────
# EXCEL PARSER
# ─────────────────────────────────────────────

def _parse_excel(fpath: Path) -> list[dict]:
    """Parse a Stock Excel export into a list of {header: value} dicts."""
    try:
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
    except Exception as e:
        print(f"  [ATG] Could not open {fpath.name}: {e}")
        return []

    ws = wb[wb.sheetnames[0]]

    # Find header row — look for recognisable Stock keywords
    headers = None
    header_row_idx = 0
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        joined = " ".join(cells)
        if any(k in joined for k in ("tank", "volume", "level", "product")):
            headers = cells
            header_row_idx = row_idx
            break

    if not headers:
        wb.close()
        print(f"  [ATG] Header row not found in {fpath.name}")
        return []

    rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        row_data = {
            headers[i]: (str(v).strip() if v is not None else "")
            for i, v in enumerate(row)
            if i < len(headers)
        }
        if any(row_data.values()):
            rows.append(row_data)

    wb.close()
    return rows


# ─────────────────────────────────────────────
# ROW PARSER
# ─────────────────────────────────────────────

def _find_col(row: dict, *keywords) -> str | None:
    """Return the value of the first key containing all keywords (case-insensitive)."""
    kw = [k.lower() for k in keywords]
    for key, val in row.items():
        if all(k in key for k in kw):
            return val
    return None


def _safe_float(val: str | None) -> float | None:
    if val is None:
        return None
    try:
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _parse_row(row: dict) -> dict | None:
    """
    Convert one raw grid/Excel row into a structured ATG reading dict.

    Returns None if the row cannot be mapped to a known product.
    The IRAS Stock tab column names are not guaranteed — we match by keyword
    substring so that "Product Code", "Prod Code", "Product" all resolve.
    """
    # Product — try "product code" first, then bare "product"
    product_raw = _find_col(row, "product", "code") or _find_col(row, "product")
    if not product_raw:
        return None
    product = product_raw.strip().upper()
    if product not in PRODUCT_TANK_MAP:
        return None

    tank_info = PRODUCT_TANK_MAP[product]

    # IRAS Stock tab column names (confirmed from live portal):
    #   product dip  → level in mm
    #   net qty      → net volume in litres (product qty minus water qty)
    #   stock date   → DD-MM-YYYY
    #   stock time   → HH:MM:SS
    level_mm      = _safe_float(_find_col(row, "product dip") or _find_col(row, "dip"))
    volume_litres = _safe_float(
        _find_col(row, "net qty") or _find_col(row, "product qty") or _find_col(row, "volume")
    )
    capacity_litres = tank_info["capacity_litres"]  # no capacity column in grid

    # pct_full computed from volume/capacity
    pct_full = None
    if volume_litres is not None and capacity_litres > 0:
        pct_full = round(volume_litres / capacity_litres * 100, 2)

    # scraped_at: combine stock date + stock time columns
    date_raw = _find_col(row, "stock date") or _find_col(row, "date")
    time_raw = _find_col(row, "stock time") or _find_col(row, "time")
    scraped_at = None
    if date_raw and time_raw:
        ts_raw = f"{date_raw.strip()} {time_raw.strip()}"
        for fmt in ("%d-%m-%Y %H:%M:%S", "%d/%m/%Y %H:%M:%S",
                    "%Y-%m-%d %H:%M:%S", "%d-%m-%Y %H:%M"):
            try:
                scraped_at = datetime.strptime(ts_raw, fmt)
                break
            except ValueError:
                continue
    if scraped_at is None:
        scraped_at = datetime.utcnow()

    return {
        "tank_id":        tank_info["tank_id"],
        "product":        product,
        "level_mm":       level_mm,
        "volume_litres":  volume_litres,
        "capacity_litres": capacity_litres,
        "pct_full":       pct_full,
        "is_reliable":    tank_info["is_reliable"],
        "scraped_at":     scraped_at,
    }


# ─────────────────────────────────────────────
# DB WRITE
# ─────────────────────────────────────────────

def save_readings_to_db(readings: list[dict]) -> int:
    """
    Upsert ATG readings into `tank_readings`.

    Uses the UniqueConstraint (scraped_at, tank_id) — skips if a row already
    exists for that snapshot time + tank to keep reruns idempotent.
    Returns the count of rows inserted.
    """
    if not readings:
        print("  [db] No ATG readings to save.")
        return 0

    try:
        import sys as _sys
        _project_root = str(Path(__file__).parent.parent)
        if _project_root not in _sys.path:
            _sys.path.insert(0, _project_root)
        from pumpvision import create_app
        from pumpvision.models import db, TankReading

        app = create_app()
        with app.app_context():
            saved = 0
            for r in readings:
                existing = db.session.query(TankReading).filter_by(
                    scraped_at=r["scraped_at"],
                    tank_id=r["tank_id"],
                ).first()

                if existing:
                    print(f"  [db] Skip duplicate: tank {r['tank_id']} @ {r['scraped_at']}")
                    continue

                db.session.add(TankReading(
                    scraped_at=r["scraped_at"],
                    tank_id=r["tank_id"],
                    product=r["product"],
                    level_mm=r["level_mm"],
                    volume_litres=r["volume_litres"],
                    capacity_litres=r["capacity_litres"],
                    pct_full=r["pct_full"],
                    is_reliable=r["is_reliable"],
                ))
                saved += 1

            db.session.commit()
            print(f"  [db] Saved {saved} ATG reading(s) to tank_readings")
            return saved

    except Exception as e:
        print(f"  [db] ERROR saving ATG readings: {e}")
        return 0


# ─────────────────────────────────────────────
# MAIN ENTRY (called from daily_scrape.py)
# ─────────────────────────────────────────────

async def run_atg(page, output_dir: Path | None = None, dry_run: bool = False) -> list[dict]:
    """
    Navigate to the Stock tab, scrape the current ATG snapshot, and write to DB.

    Called from daily_scrape.py with an already-authenticated IRAS page.
    Returns the list of parsed reading dicts (empty on failure).
    """
    print(f"\n{'='*55}")
    print(f"  JOB 5 — ATG Stock Snapshot  (current reading, not date-specific)")
    print(f"{'='*55}")

    if output_dir is None:
        output_dir = Path(os.environ.get("OUTPUT_FOLDER", r"C:\IRAS_Data")) / "ATG"
    output_dir.mkdir(parents=True, exist_ok=True)

    await navigate_to_stock(page)

    # Click Show to load the table (Stock tab may not need date inputs)
    try:
        show_btn = page.locator("button:has-text('Show')").first
        await show_btn.wait_for(state="visible", timeout=5_000)
        await show_btn.click()
        await page.wait_for_timeout(2000)
    except Exception:
        print("  [ATG] No Show button — table may auto-load")
        await page.wait_for_timeout(2000)

    # Wait for rows
    try:
        await page.wait_for_selector(
            ".ag-row, .ag-overlay-no-rows-wrapper",
            timeout=TABLE_LOAD_TIMEOUT * 1000,
        )
    except PlaywrightTimeout:
        print("  [ATG] Table load timeout")

    row_count = await page.locator(".ag-row").count()
    if row_count == 0:
        print("  [ATG] No rows in Stock table — skipping")
        return []

    print(f"  [ATG] {row_count} row(s) in Stock table")

    # Strategy 1: read ag-Grid cells directly
    raw_rows = await _read_ag_grid(page)

    # Strategy 2: fallback to Excel download
    if not raw_rows:
        print("  [ATG] Grid read empty — trying Excel download")
        now = datetime.now()
        fname = f"ATG_Stock_{now.strftime('%Y%m%d_%H%M')}.xlsx"
        fpath = output_dir / fname
        try:
            async with page.expect_download(timeout=DOWNLOAD_TIMEOUT * 1000) as dl_info:
                await page.locator("button.export-excel-button").click()
            download = await dl_info.value
            await download.save_as(str(fpath))
            print(f"  [ATG] Downloaded: {fname}")
            raw_rows = _parse_excel(fpath)
        except Exception as e:
            print(f"  [ATG] Excel download failed: {e}")

    if not raw_rows:
        print("  [ATG] Could not read Stock table by any method")
        return []

    # Parse rows into structured readings
    readings = []
    for row in raw_rows:
        parsed = _parse_row(row)
        if parsed:
            readings.append(parsed)
        else:
            # Log unrecognised rows for debugging
            product_hint = _find_col(row, "product") or "?"
            print(f"  [ATG] Skipped row (product={product_hint!r}): {list(row.keys())[:6]}")

    if raw_rows:
        print(f"  [ATG] Column names: {list(raw_rows[0].keys())}")
    print(f"  [ATG] Parsed {len(readings)}/{len(raw_rows)} rows → products: "
          f"{[r['product'] for r in readings]}")

    if readings:
        for r in readings:
            vol = f"{r['volume_litres']:.0f} L" if r["volume_litres"] is not None else "—"
            pct = f"{r['pct_full']:.1f}%" if r["pct_full"] is not None else "—"
            print(f"    Tank {r['tank_id']} ({r['product']}): {vol}  {pct}  "
                  f"@ {r['scraped_at'].strftime('%H:%M')}")

        if dry_run:
            print(f"  [dry-run] DB write skipped — would have saved {len(readings)} ATG reading(s)")
        else:
            save_readings_to_db(readings)

    return readings


# ─────────────────────────────────────────────
# AUTONOMOUS LOGIN (standalone use)
# ─────────────────────────────────────────────

_CAPTCHA_PROMPT = (
    "Read the characters in this CAPTCHA image exactly as they appear. "
    "Reply with only the characters, no spaces, no punctuation, nothing else. "
    "Ignore any strikethrough or diagonal lines across the text."
)

_MAX_LOGIN_ATTEMPTS = 3


def _solve_captcha(image_bytes: bytes, api_key: str) -> str:
    import anthropic
    client = anthropic.Anthropic(api_key=api_key)
    b64 = base64.standard_b64encode(image_bytes).decode("utf-8")
    msg = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=64,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image",
                 "source": {"type": "base64", "media_type": "image/png", "data": b64}},
                {"type": "text", "text": _CAPTCHA_PROMPT},
            ],
        }],
    )
    return msg.content[0].text.strip()


async def _autonomous_login(page, username: str, password: str, api_key: str) -> bool:
    login_url = IRAS_URL
    print(f"\n[login] Autonomous login — up to {_MAX_LOGIN_ATTEMPTS} attempts")

    for attempt in range(1, _MAX_LOGIN_ATTEMPTS + 1):
        print(f"  [login] Attempt {attempt}/{_MAX_LOGIN_ATTEMPTS}")

        if attempt > 1:
            refresh_selectors = [
                "img[src*='refresh']", "img[src*='reload']",
                "a[onclick*='captcha']", ".captcha-refresh", "#captchaRefresh",
            ]
            refreshed = False
            for sel in refresh_selectors:
                loc = page.locator(sel).first
                try:
                    if await loc.count() > 0:
                        await loc.click()
                        await page.wait_for_timeout(1000)
                        refreshed = True
                        break
                except Exception:
                    continue
            if not refreshed:
                # Raw error message suppressed — may contain proxy host/port if connection dropped.
                try:
                    await page.goto(login_url, wait_until="networkidle", timeout=30_000)
                except PlaywrightTimeout:
                    pass  # timeout on retry is non-critical; continue attempt
                except Exception as _retry_nav_exc:
                    print(f"  [login] Retry navigation failed: {safe_exc_name(_retry_nav_exc)}")
                await page.wait_for_timeout(800)

        # Screenshot CAPTCHA
        captcha_img = None
        for sel in ["img[src*='captcha']", "img[src*='Captcha']", "img[src*='kaptcha']",
                    "img[id*='captcha']", "img[class*='captcha']", "img[alt*='aptcha']",
                    "form img"]:
            loc = page.locator(sel).first
            try:
                if await loc.count() > 0 and await loc.is_visible(timeout=1500):
                    captcha_img = loc
                    break
            except Exception:
                continue

        if captcha_img is None:
            print("  [login] CAPTCHA image not found")
            continue

        img_bytes = await captcha_img.screenshot()
        captcha_text = _solve_captcha(img_bytes, api_key)
        print(f"  [login] CAPTCHA solved: {captcha_text}")

        await page.wait_for_timeout(800)

        # Role dropdown — try native select first, then MUI combobox
        try:
            native = page.locator("select").first
            if await native.count() > 0 and await native.is_visible(timeout=2000):
                await native.select_option(label="Dealer")
            else:
                combo = page.locator("div[role='combobox'], .MuiSelect-select").first
                if await combo.count() > 0 and await combo.is_visible(timeout=2000):
                    await combo.click()
                    await page.wait_for_timeout(500)
                    dealer = page.locator("li[role='option']:has-text('Dealer')").first
                    await dealer.wait_for(state="visible", timeout=3000)
                    await dealer.click()
        except Exception:
            pass

        await page.wait_for_timeout(600)

        # Username
        for sel in ["input[name='username']", "input[name='userId']",
                    "input[placeholder*='Username']", "input[placeholder*='User']"]:
            loc = page.locator(sel).first
            try:
                if await loc.count() > 0 and await loc.is_visible(timeout=1000):
                    await loc.fill(username)
                    break
            except Exception:
                continue

        # Password
        try:
            pw = page.locator("input[type='password']").first
            if await pw.count() > 0 and await pw.is_visible(timeout=2000):
                await pw.fill(password)
        except Exception:
            pass

        # CAPTCHA input
        cap_input = None
        for sel in ["input[name*='captcha']", "input[name*='Captcha']",
                    "input[id*='captcha']", "input[placeholder*='aptcha']"]:
            loc = page.locator(sel).first
            try:
                if await loc.count() > 0 and await loc.is_visible(timeout=1000):
                    cap_input = loc
                    break
            except Exception:
                continue
        if cap_input is None:
            inputs = page.locator("input[type='text']:visible")
            count = await inputs.count()
            if count > 0:
                cap_input = inputs.nth(count - 1)
        if cap_input:
            await cap_input.fill(captcha_text)

        # Submit
        submitted = False
        for sel in ["button[type='submit']", "input[type='submit']",
                    "button:has-text('Login')", "button:has-text('Sign In')"]:
            loc = page.locator(sel).first
            try:
                if await loc.count() > 0 and await loc.is_visible(timeout=1000):
                    await loc.click()
                    submitted = True
                    break
            except Exception:
                continue
        if not submitted:
            await page.keyboard.press("Enter")

        await page.wait_for_timeout(3000)

        try:
            await page.wait_for_function(
                "() => !window.location.href.includes('/login')", timeout=5000)
            print("  [login] SUCCESS")
            await page.wait_for_timeout(2000)
            return True
        except PlaywrightTimeout:
            pass

        if "/login" not in page.url:
            print("  [login] SUCCESS")
            await page.wait_for_timeout(2000)
            return True

        print(f"  [login] Failed — CAPTCHA was: {captcha_text}")

    print(f"[login] FAILED after {_MAX_LOGIN_ATTEMPTS} attempts")
    return False


# ─────────────────────────────────────────────
# STANDALONE ENTRY POINT
# ─────────────────────────────────────────────

async def _main_standalone():
    """Run ATG scraper standalone with autonomous CAPTCHA login."""
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent.parent / ".env")

    username  = os.environ.get("IRAS_USERNAME", "")
    password  = os.environ.get("IRAS_PASSWORD", "")
    api_key   = os.environ.get("ANTHROPIC_API_KEY", "")
    output_dir = Path(os.environ.get("OUTPUT_FOLDER", r"C:\IRAS_Data")) / "ATG"
    output_dir.mkdir(parents=True, exist_ok=True)

    print()
    print("IRAS ATG Stock Exporter (standalone)")
    print(f"Output: {output_dir}")
    print()

    if not all([username, password, api_key]):
        missing = [k for k, v in [("IRAS_USERNAME", username),
                                   ("IRAS_PASSWORD", password),
                                   ("ANTHROPIC_API_KEY", api_key)] if not v]
        print(f"ERROR: missing env vars: {missing}")
        return

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=True,
            args=["--no-sandbox", "--disable-dev-shm-usage"],
        )
        _iras_proxy = iras_proxy_cfg()
        _ctx_kw: dict = {
            "accept_downloads": True,
            "viewport": {"width": 1400, "height": 900},
            "user_agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        }
        if _iras_proxy is not None:
            _ctx_kw["proxy"] = _iras_proxy
        # Context/page setup — exceptions suppressed to avoid leaking proxy config.
        try:
            context = await browser.new_context(**_ctx_kw)
            page = await context.new_page()
        except Exception as _setup_exc:
            print(f"  [IRAS] Browser/context setup failed: {safe_exc_name(_setup_exc)}")
            await browser.close()
            return
        print(f"  [IRAS] proxy : {'yes' if IRAS_PROXY_ENABLED else 'no'}")
        print(f"[step 0] Loading: {IRAS_URL}")
        # Initial navigation — raw error message suppressed (may contain proxy host/port).
        try:
            await page.goto(IRAS_URL, wait_until="networkidle", timeout=30_000)
        except PlaywrightTimeout:
            print(f"  [IRAS] Navigation timeout (networkidle) — continuing")
        except Exception as _nav_exc:
            print(f"  [IRAS] Initial navigation failed: {safe_exc_name(_nav_exc)}")
            await browser.close()
            return
        await page.wait_for_timeout(1000)

        if not await _autonomous_login(page, username, password, api_key):
            print("\nABORTED — login failed.")
            await browser.close()
            return

        await run_atg(page, output_dir)

        print("\n[DONE]")
        await browser.close()


if __name__ == "__main__":
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent.parent / ".env")
    asyncio.run(_main_standalone())
